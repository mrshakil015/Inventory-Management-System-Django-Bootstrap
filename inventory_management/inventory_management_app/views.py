from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.core.files.storage import default_storage
from django.contrib.auth.decorators import login_required
from .models import *
from .forms import *
import random
import string
from django.db.models import Sum, Count, F
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from decimal import Decimal,InvalidOperation
from django.db.models.functions import TruncDate
from django.utils.timezone import now
from datetime import timedelta, datetime
from django.http import HttpResponse,JsonResponse
import csv
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.hashers import check_password
from weasyprint import HTML
from django.template.loader import render_to_string
import io
import os
import pandas as pd
from .decorators import user_has_access
from urllib.parse import quote
from django.core.mail import EmailMessage
from django.db.models import Q
import openpyxl
import math


def user_login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        try:
            user = InventoryUser.objects.get(username=username)
            user = authenticate(request, username=username, password=password)
            
            if user is not None and user.is_active:
                login(request, user)
                request.session.cycle_key()  # Regenerate the session ID for security
                return redirect('dashboard')
            else:
                messages.warning(request, "Invalid username or password.")
        except InventoryUser.DoesNotExist:
            messages.warning(request, "User does not exist.")

    return render(request, "auth/login.html")

@login_required
def user_logout(request):
    logout(request)
    return redirect('user_login')

@login_required
def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        user = request.user

        if not check_password(old_password, user.password):
            messages.error(request, "Old password is incorrect.")
            return redirect('change_password')
        
        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect('change_password')

        user.set_password(new_password)
        user.save()

        # Keep the user logged in after password change
        update_session_auth_hash(request, user)

        messages.success(request, "Password changed successfully!")
        return redirect('dashboard')

    return render(request, 'auth/change-password.html')

@login_required
def dashboard(request):
    total_employees = EmployeeModel.objects.count()
    total_customers = CustomerModel.objects.count()
    
    medicine_query = MedicineModel.objects.annotate(
        total_quantity_value=F('total_quantity') * F('unit_sale_price')  
    ).aggregate(
        total_unit_medicine=Sum('total_medicine'),                    
        total_quantity=Sum('total_quantity'),                  
        current_product_value=Sum('total_quantity_value'),      
        total_medicine_count=Count('id')                         
    )
    
    # Handle None values safely
    current_product_value = medicine_query['current_product_value'] or 0
    formatted_current_product_value = Decimal(str(current_product_value)).quantize(Decimal('0.001'))

    total_quantity = medicine_query['total_quantity'] or 0
    rounded_total_quantity = math.ceil(float(total_quantity))
    formatted_total_quantity = Decimal(rounded_total_quantity).quantize(Decimal('1'))

    # Update the dictionary with the formatted value
    medicine_query['current_product_value'] = formatted_current_product_value
    medicine_query['total_quantity'] = formatted_total_quantity
    
    medicine_data = MedicineModel.objects.all()[:6]
    billing_query = BillingModel.objects.annotate(
        total_sale_amount_value=F('total_amount')
    ).aggregate(
        total_billing=Count('id'),
        total_sale_amount=Sum('total_sale_amount_value'),
        
    )

    
    total_purchase_amount = MedicineStockModel.objects.aggregate(total_amount=Sum('total_amount'))
    total_wastage_quantity = BottleBreakageModel.objects.aggregate(total_wastage_quantity=Sum('lost_quantity'))

    thirty_days_ago = now().date() - timedelta(days=30)
    
    latest_billing= BillingModel.objects.all().order_by('-id')[:5]
    latest_stock = MedicineStockModel.objects.all().order_by('-id')[:5]
    
    billings = (
        BillingModel.objects
        .filter(billing_date__date__gte=thirty_days_ago) 
        .annotate(billing_date_only=TruncDate('billing_date'))
        .values('billing_date_only')
        .annotate(total_amount=Sum('total_amount'))
        .order_by('billing_date_only')
    )

    # Convert data for ApexCharts
    billing_chart_data = {
        "dates": [billing['billing_date_only'].strftime('%Y-%m-%d') for billing in billings],
        "totals": [float(billing['total_amount']) for billing in billings]
    }

    context = {
        "total_employees": total_employees,
        "total_customers": total_customers,
        "total_billing": billing_query['total_billing'],
        "total_medicine_count": medicine_query['total_medicine_count'] or 0,
        "total_unit_medicine": medicine_query['total_unit_medicine'] or 0,
        "total_medicine_pack": medicine_query['total_quantity'] or 0,
        "current_product_value": medicine_query['current_product_value'] or 0,
        "total_purchase_amount": total_purchase_amount['total_amount'] or 0,
        "total_sale_amount": billing_query['total_sale_amount'] or 0,
        "total_wastage_quantity": total_wastage_quantity['total_wastage_quantity'] or 0,
        "total_revenue_amount": 0,
        "billing_chart_data": billing_chart_data,  
        "latest_billing":latest_billing,
        "latest_stock":latest_stock,
        "medicine_data":medicine_data,
    }

    return render(request, "index.html", context)

@login_required
@user_has_access('employee_view','employee_management')
def employee_list(request):
    employees = EmployeeModel.objects.all()
    context = {
        'employees': employees
    }
    return render(request, 'employees/employee-list.html', context)

@user_has_access('employee_management')
@login_required
def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            employee = form.save(commit=False)
            user_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
            username = f"EID-{random.randint(100000, 999999)}"

            while InventoryUser.objects.filter(username=username).exists():
                username = f"EID-{random.randint(100000, 999999)}"

            user = InventoryUser.objects.create(
                username=username,
                email=form.cleaned_data['email'],
                role=form.cleaned_data['role'],
            )
            user.set_password(user_password)
            
            # If role is 'Admin', store predefined access list
            if user.role == 'Admin':
                user_access_list = ['employee_management', 'employee_view', 'customer_management', 'customer_view', 'product_management', 'product_view', 'low_stocks', 'billing_management', 'inventory_report', 'wastage_report', 'billing_report']

            else:
                user_access_list = form.cleaned_data['user_access_list']

            user.user_access_list = ','.join(user_access_list)
            
            user.save()

            employee.employee_id = username
            employee.created_by = request.user
            employee.employee_user = user
            employee.save()

            send_mail(
                subject="Account Login Info",
                message=f"Your Account successfully created. Username: {username}, Password: {user_password}",
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[user.email],
            )

            messages.success(request, "Employee added successfully!")
            return redirect('employee_list')
    else:
        form = EmployeeForm()

    return render(request, 'employees/add-employee.html', {'form': form})


@user_has_access('employee_management')
@login_required
def update_employee(request, pk):
    employee = get_object_or_404(EmployeeModel, pk=pk)
    old_image_path = employee.employee_picture.path if employee.employee_picture else None
    
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            if 'employee_picture' in request.FILES and old_image_path:
                if default_storage.exists(old_image_path):
                    default_storage.delete(old_image_path)
            form.save()
            messages.success(request, "Employee updated successfully!")
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'employees/update-employee.html', {'form': form})

@login_required
@user_has_access('employee_management')
def delete_employee(request, pk):
    employee = get_object_or_404(EmployeeModel, pk=pk)
    image_path = employee.employee_picture.path if employee.employee_picture else None

    if image_path and default_storage.exists(image_path):
        default_storage.delete(image_path)
    employee.delete()
    messages.success(request, "Employee deleted successfully!")
    return redirect('employee_list')

def delete_selected_employee(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_employee")

        if selected_ids:
            employees = EmployeeModel.objects.filter(id__in=selected_ids)

            for employee in employees:
                image_path = employee.employee_picture.path if employee.employee_picture else None

                if image_path and default_storage.exists(image_path):
                    default_storage.delete(image_path)

                employee.delete()

            messages.success(request, "Selected employees deleted successfully.")
        else:
            messages.warning(request, "No employees selected for deletion.")

    return redirect("employee_list")
            
# --------Customer Functionalities
@login_required
@user_has_access('customer_management','billing_management')
def add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            phone = customer.customer_phone
            email = customer.customer_email
            
            # Check if phone number already exists
            if CustomerModel.objects.filter(customer_phone=phone).exists():
                messages.warning(request, "Phone number is already taken!")
                return render(request, 'customers/add-customer.html', {'form': form})
            
            # Check if email already exists
            if CustomerModel.objects.filter(customer_email=email).exists():
                messages.warning(request, "Email is already taken!")
                return render(request, 'customers/add-customer.html', {'form': form})

            # If no errors, save the customer
            customer.created_by = request.user
            customer.save()
            messages.success(request, "Customer added successfully!")
            return redirect('customer_list')
    else:
        form = CustomerForm()
    
    return render(request, 'customers/add-customer.html', {'form': form})


# List Customers
@login_required
@user_has_access('customer_management','customer_view','billing_management')
def customer_list(request):
    customers = CustomerModel.objects.all()
    context = {
        'customers': customers
    }
    return render(request, 'customers/customer-list.html', context)

# Update Customer
@login_required
@user_has_access('customer_management','billing_management')
def update_customer(request, customer_id):
    customer = get_object_or_404(CustomerModel, id=customer_id)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
        
    context = {
        'form':form,
    }
    
    return render(request, 'customers/update-customer.html', context)

# Delete Customer
@login_required
@user_has_access('customer_management','billing_management')
def delete_customer(request, customer_id):
    customer = get_object_or_404(CustomerModel, id=customer_id)
    customer.delete()
    return redirect('customer_list')

@login_required
@user_has_access('customer_management','billing_management')
def delete_selected_customers(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_customers")
        if selected_ids:
            CustomerModel.objects.filter(id__in =selected_ids).delete()
            messages.success(request, "Selected customer deleted successfully.")
        else:
            messages.warning(request, "No customer selected for deletion.")
    return redirect('customer_list')

#---Medicine Category
@login_required
@user_has_access('product_management')
def medicine_category_list(request):
    medicine_category = MedicineCategoryModel.objects.all()
    context = {
        'medicine_category':medicine_category
    }
    return render(request,'medicine_category/medicine-category-list.html',context)


@login_required
@user_has_access('product_management')
def add_medicine_category(request):
    if request.method == 'POST':
        form = MedicineCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('medicine_category_list')
    else:
        form = MedicineCategoryForm()
    context = {
        'form':form
    }
    return render(request,'medicine_category/add-medicine-category.html',context)

@login_required
@user_has_access('product_management')
def update_medicine_category(request,pk):
    category = MedicineCategoryModel.objects.get(id=pk)
    if request.method == 'POST':
        form = MedicineCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('medicine_category_list')
    else:
        form = MedicineCategoryForm(instance=category)
    context = {
        'form':form
    }
    return render(request,'medicine_category/update-medicine-category.html',context)

@login_required
@user_has_access('product_management')
def delete_medicine_category(request, pk):
    category = MedicineCategoryModel.objects.get(id=pk)
    category.delete()
    return redirect('medicine_category_list')

def delete_selected_medicine_categories(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_medicine_categories")
        if selected_ids:
            MedicineCategoryModel.objects.filter(id__in=selected_ids).delete()
            messages.success(request, "Selected medicine category deleted successfully")
        else:
            messages.warning(request, "No medicine category selected for deletion.")
    return redirect('medicine_category_list')

#---Medicine Unit
@login_required
@user_has_access('product_management')
def medicine_unit_list(request):
    medicine_unit = MedicineUnitModel.objects.all()
    context = {
        'medicine_unit':medicine_unit
    }
    return render(request,'medicine_unit/medicine-unit-list.html',context)


@login_required
@user_has_access('product_management')
def add_medicine_unit(request):
    if request.method == 'POST':
        form = MedicineUnitForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Unit added successfully!")
            return redirect('medicine_unit_list')
    else:
        form = MedicineUnitForm()
    context = {
        'form':form
    }
    return render(request,'medicine_unit/add-medicine-unit.html',context)

@login_required
@user_has_access('product_management')
def update_medicine_unit(request,pk):
    unit = MedicineUnitModel.objects.get(id=pk)
    if request.method == 'POST':
        form = MedicineUnitForm(request.POST, instance=unit)
        if form.is_valid():
            form.save()
            messages.success(request, "Unit updated successfully!")
            return redirect('medicine_unit_list')
    else:
        form = MedicineUnitForm(instance=unit)
    context = {
        'form':form
    }
    return render(request,'medicine_unit/update-medicine-unit.html',context)

@login_required
@user_has_access('product_management')
def delete_medicine_unit(request, pk):
    unit = MedicineUnitModel.objects.get(id=pk)
    unit.delete()
    messages.success(request, "Unit deleted successfully!")
    return redirect('medicine_unit_list')

@login_required
@user_has_access('product_management')
def delete_selected_medicine_units(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_units")
        if selected_ids:
            MedicineUnitModel.objects.filter(id__in=selected_ids).delete()
            messages.success(request, "Selected medicine units deleted successfully.")
        else:
            messages.warning(request, "No units selected for deletion.")
    return redirect("medicine_unit_list")


#------Medicine
from django.core.paginator import Paginator
from django.db.models import Q
@login_required
@user_has_access('product_management','product_view','low_stocks','billing_management')
def medicine_list(request):
    per_page = request.GET.get('per_page', 10)
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10

    query = request.GET.get('q', '')

    medicine_list = MedicineModel.objects.all()

    if query:
        medicine_list = medicine_list.filter(
            Q(medicine_name__icontains=query) |
            Q(sku__icontains=query) |
            Q(batch_number__icontains=query)
        )

    paginator = Paginator(medicine_list, per_page)
    page_number = request.GET.get('page')
    medicines = paginator.get_page(page_number)
    
    if request.GET.get('download') == 'true':
        # Create an Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Medicine Data"

        # Write headers
        sheet.append(['Batch No','Medicine Name','Brand Name', 'Medicine Category', 'Medicine Type','Unit Sale Price','GST (%)','Pack Size', 'Total Case Pack', 'Total Medicine','Status'])
        
        # Write medicine data with two empty columns
        for medicine in medicines:
            pack_size = str(medicine.pack_size) + " " + str(medicine.pack_units.unit_name) 
            total_medicine = str(medicine.total_medicine) + " " + str(medicine.pack_units.unit_name)

            sheet.append([
                medicine.batch_number or "",
                medicine.medicine_name or "",
                medicine.brand_name or "",
                medicine.medicine_category.category_name if medicine.medicine_category else "",
                medicine.medicine_type or "",
                medicine.unit_sale_price if medicine.unit_sale_price is not None else "",
                medicine.gst_percentage if medicine.gst_percentage is not None else "",
                pack_size,
                medicine.total_quantity if medicine.total_quantity is not None else "",
                total_medicine,
                medicine.stocks or ""
            ])

        # Prepare HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="all_medicine_data.xlsx"'

        # Save workbook to response
        workbook.save(response)
        return response
    
    if request.GET.get('download-format') == 'true':
        # Create an Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Medicine format sheet"

        # Write headers
        sheet.append(['batch_number','medicine_name','brand_name','medicine_category','medicine_type','pack_size','pack_units','unit_sale_price','gst_percentage','purchase_price','total_quantity','description'])

        for medicine in medicines:
            sheet.append(['', '', '','', '', '','', '', '','', '','', ''])

        # Prepare HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="medicine_format_sheet.xlsx"'

        # Save workbook to response
        workbook.save(response)
        return response
    return render(request, 'medicines/medicine-list.html', {
        'medicines': medicines,
        'per_page': per_page,
        'query': query
    })

def sku_generate():
    return f"MED{random.randint(10000, 99999)}"

@login_required
@user_has_access('product_management')
def add_medicine(request):
    if request.method == 'POST':
        form = MedicineForm(request.POST, request.FILES)
        total_quantity = form.data.get('total_quantity')
        purchase_price = form.data.get('purchase_price')
        if form.is_valid():
            medicine = form.save(commit=False)
            medicine.created_by = request.user
            while True:
                sku_no = sku_generate()
                if not MedicineModel.objects.filter(sku=sku_no).exists():
                    break
            
            medicine.sku = sku_no
            # Remove extra spaces from each part
            medicine_name = " ".join(medicine.medicine_name.split())  # Removes multiple spaces
            pack_units = " ".join(medicine.pack_units.unit_name.split())
            pack_size = str(medicine.pack_size).strip()
            
            # Construct the final name
            full_medicine_name = f"{medicine_name} {pack_size} {pack_units}"

            # Case-insensitive check without modifying actual stored name
            if MedicineModel.objects.filter(medicine_name__iexact=full_medicine_name).exists():
                messages.error(request, "This medicine already exists!")
            else:
                medicine.medicine_name = full_medicine_name
                medicine.save()
                if total_quantity or purchase_price:
                    stock = MedicineStockModel.objects.create(
                        medicine= medicine,
                        total_quantity=int(total_quantity) if total_quantity else 0,
                        purchase_price=Decimal(purchase_price) if purchase_price else 0,
                        created_by=request.user,
                        
                    )
                    stock.save()
                    # Update medicine stock
                    medicine = stock.medicine
                    medicine.total_quantity += stock.total_quantity
                    medicine.save()
                messages.success(request, "Medicine added successfully!")
                return redirect('medicine_list')
        else:
            messages.warning(request, "Error in form submission. Please try again.")
    else:
        form = MedicineForm()
    
    return render(request, 'medicines/add-medicine.html', {'form': form})

@login_required
@user_has_access('product_management')
def update_medicine(request, pk):
    medicine = get_object_or_404(MedicineModel, pk=pk)
    
    if request.method == 'POST':
        form = MedicineForm(request.POST, request.FILES, instance=medicine)
        if form.is_valid():

            form.save()
            messages.success(request, "Medicine updated successfully!")
            return redirect('medicine_list')  
        else:
            messages.warning(request, "Error in form submission. Please try again.")
    else:
        form = MedicineForm(instance=medicine)
    
    return render(request, 'medicines/update-medicine.html', {'form': form})

@login_required
@user_has_access('product_management')
def delete_medicine(request, pk):
    medicine = get_object_or_404(MedicineModel, pk=pk)
    
    # Delete the medicine and its associated image
    if medicine.medicine_picture:
        try:
            medicine.medicine_picture.delete()
        except Exception as e:
            pass
    
    medicine.delete()
    messages.success(request, "Medicine deleted successfully!")
    return redirect('medicine_list')

@login_required
@user_has_access('product_management')
def delete_selected_medicines(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_medicines")
        if selected_ids:
            MedicineModel.objects.filter(id__in=selected_ids).delete()
            messages.success(request, "Selected medicine deleted successfully.")
        else:
            messages.warning(request, "No medicine selected for deletion.")
    return redirect("medicine_list")

def medicine_detail(request, pk):
    medicine = get_object_or_404(MedicineModel, pk=pk)
    stocks = MedicineStockModel.objects.filter(medicine=pk).order_by('-id')
    return render(request, 'medicines/medicine_detail.html', {'medicine': medicine, 'stocks': stocks})

#-----Medicine Upload to Excel
@login_required
@user_has_access('product_management')
def upload_medicine(request):
    if request.method == "POST":
        try:
            # Step 1: Read the uploaded file
            excel_file = request.FILES["file"]
            file_name = excel_file.name

            if file_name.endswith(".csv"):
                df = pd.read_csv(excel_file)
            elif file_name.endswith(".xls") or file_name.endswith(".xlsx"):
                df = pd.read_excel(excel_file, engine="openpyxl")
            else:
                return JsonResponse({"message": "Unsupported file format! Please upload CSV or Excel."}, status=400)

            # Step 2: Validate columns
            df.columns = df.columns.str.strip().str.lower()
            required_columns = {
                "batch_number", "brand_name", "medicine_name", "medicine_category", 
                "medicine_type", "pack_units", "description", "pack_size", 
                "unit_sale_price", "purchase_price", "total_quantity", "gst_percentage"
            }
            
            missing_columns = required_columns - set(df.columns)
            if missing_columns:
                return JsonResponse({"message": f"Missing columns: {', '.join(list(missing_columns))}"}, status=400)

            # Step 3: Process rows
            valid_rows = []
            stock_records = []
            invalid_rows = []
            valid_flag = False

            for _, row in df.iterrows():
                row_dict = row.to_dict()
                errors = []
                if pd.isna(row["medicine_name"]) or row["medicine_name"].strip() == "":
                    errors.append("Missing medicine name")
                
                try:
                    category = MedicineCategoryModel.objects.get(category_name=row["medicine_category"])
                except MedicineCategoryModel.DoesNotExist:
                    category = None

                try:
                    unit = MedicineUnitModel.objects.get(unit_name=row["pack_units"])
                    while True:
                        sku_no = sku_generate()
                        if not MedicineModel.objects.filter(sku=sku_no).exists():
                            break

                    created_by = request.user
                    full_medicine_name = f"{row['medicine_name']} {row['pack_size']} {unit}"
                    
                except MedicineUnitModel.DoesNotExist:
                    errors.append(f"Invalid pack unit: {row['pack_units']}")
                    full_medicine_name = None

                if row["medicine_name"] and full_medicine_name:
                    if MedicineModel.objects.filter(medicine_name=full_medicine_name).exists():
                        continue

                # Process GST percentage
                if row.get("gst_percentage"):
                    gst_value = str(row["gst_percentage"]).strip()
                    if gst_value.endswith("%"):
                        gst_percentage = gst_value
                    else:
                        try:
                            gst_float = float(gst_value)
                            if gst_float > 1: 
                                gst_percentage = f"{int(gst_float)}%"
                            else: 
                                gst_percentage = f"{int(gst_float * 100)}%"
                        except (ValueError, TypeError):
                            gst_percentage = "0%"
                else:
                    gst_percentage = "0%"

                try:
                    total_quantity = Decimal(str(row["total_quantity"])) if not pd.isna(row["total_quantity"]) else Decimal('0')
                except (ValueError, TypeError, InvalidOperation):
                    total_quantity = Decimal('0')
                
                try:
                    purchase_price = Decimal(str(row["purchase_price"])) if not pd.isna(row["purchase_price"]) else Decimal('0')
                except (ValueError, TypeError, InvalidOperation):
                    purchase_price = Decimal('0')
                    
                try:
                    unit_sale_price = Decimal(str(row["unit_sale_price"])) if not pd.isna(row["unit_sale_price"]) else Decimal('0')
                except (ValueError, TypeError, InvalidOperation):
                    unit_sale_price = Decimal('0')

                if errors:
                    row_dict["error_reason"] = "; ".join(errors)
                    invalid_rows.append(row_dict)
                else:
                    medicine = MedicineModel(
                        sku=sku_no,
                        batch_number=row["batch_number"],
                        brand_name=row["brand_name"],
                        medicine_name=full_medicine_name,
                        medicine_category=category,
                        medicine_type=row["medicine_type"],
                        pack_units=unit,
                        description=row["description"],
                        pack_size=row["pack_size"],
                        unit_sale_price=unit_sale_price,
                        gst_percentage=gst_percentage,
                        total_quantity=Decimal('0'),
                        created_by=created_by,
                    )
                    valid_rows.append(medicine)
                    
                    # Create stock record if purchase_price or total_quantity exists
                    if purchase_price > Decimal('0') or total_quantity > Decimal('0'):
                       
                        stock_records.append({
                            'sku': sku_no, 
                            'total_quantity': total_quantity,
                            'purchase_price': purchase_price,
                            'created_by': created_by
                        })

            # Step 4: Save valid rows
            if valid_rows:
                valid_flag = True
                MedicineModel.objects.bulk_create(valid_rows)
                
                for stock_data in stock_records:
                    try:
                        medicine = MedicineModel.objects.get(sku=stock_data['sku'])
                        stock = MedicineStockModel.objects.create(
                            medicine=medicine,
                            total_quantity=stock_data['total_quantity'],
                            purchase_price=stock_data['purchase_price'],
                            created_by=stock_data['created_by']
                        )
                        stock.save()
                        
                        # Update the medicine's total quantity
                        medicine.total_quantity += stock_data['total_quantity']
                        medicine.save()
                        
                    except MedicineModel.DoesNotExist:
                        pass

            # Step 5: Handle invalid rows
            if invalid_rows:
                error_df = pd.DataFrame(invalid_rows)
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    error_df.to_excel(writer, index=False)

                output.seek(0)
                
                if valid_flag:
                    messages.warning(request, "Data imported successfully! But some data had errors. Please review the error file.")
                else:
                    messages.error(request, "No valid data found. Please check the downloaded error file.")

                response = HttpResponse(
                    output.getvalue(), 
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = 'attachment; filename="error_data.xlsx"'
                return response

            return JsonResponse({"message": "Data imported successfully!"})

        except Exception as e:
            return JsonResponse({"message": f"Error importing data: {str(e)}"}, status=500)

    return JsonResponse({"message": "Invalid request method."}, status=400)


#--------Medicine Stock
@login_required
@user_has_access('product_management')
def medicine_stock_list(request):
    stocks = MedicineStockModel.objects.all()
    medicines = MedicineModel.objects.all()
    if request.GET.get('download') == 'true':
        # Create an Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Medicine list for stock"

        # Write headers
        sheet.append(['medicine_name', 'total_quantity', 'purchase_price'])

        # Write medicine data with two empty columns
        for medicine in medicines:
            sheet.append([medicine.medicine_name, '', ''])  # Adding two blank columns

        # Prepare HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="medicine_list_for_stock.xlsx"'

        # Save workbook to response
        workbook.save(response)
        return response
    
    return render(request, 'medicine_stock/medicine-stock-list.html', {'stocks': stocks})

@login_required
@user_has_access('product_management')
def add_medicine_stock(request):
    if request.method == "POST":
        form = MedicineStockForm(request.POST)
        if form.is_valid():
            stock = form.save(commit=False)
            stock.created_by = request.user
            stock.save()
            
            # Update medicine stock
            medicine = stock.medicine
            medicine.total_quantity += stock.total_quantity
            medicine.save()
            
            messages.success(request, "Medicine stock added successfully!")
            return redirect('medicine_stock_list')
    else:
        form = MedicineStockForm()
    return render(request, 'medicine_stock/add-medicine-stock.html', {'form': form})

def upload_medicine_stock(request):
    if request.method == "POST":
        form = MedicineStockUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES.get('file')
            try:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                
                # Read rows and insert into the database
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                    medicine_name, total_quantity, purchase_price = row
                    
                    if not medicine_name or total_quantity is None or total_quantity == 0 or purchase_price is None or purchase_price == 0:
                        continue  # Skip invalid rows

                    # Fetch the medicine object
                    medicine = MedicineModel.objects.filter(medicine_name=medicine_name).first()
                    if medicine:
                        # Create a stock entry only if total_quantity is valid
                        stock = MedicineStockModel.objects.create(
                            medicine=medicine,
                            total_quantity=int(total_quantity),
                            purchase_price=float(purchase_price) if purchase_price else 0,
                            created_by=request.user
                        )
                        
                        # Update the medicine stock
                        medicine.total_quantity = (medicine.total_quantity or 0) + stock.total_quantity
                        medicine.save()
                
                messages.success(request, "Stock uploaded successfully!")
                return redirect('medicine_stock_list')

            except Exception as e:
                messages.error(request, f"Error processing file: {e}")

    messages.error(request, "Invalid form submission.")
    return redirect('medicine_stock_list')

@login_required
@user_has_access('product_management')
def update_medicine_stock(request, pk):
    stock = get_object_or_404(MedicineStockModel, pk=pk)
    old_total_quantity = stock.total_quantity

    if request.method == "POST":
        form = MedicineStockForm(request.POST, instance=stock)
        if form.is_valid():
            stock = form.save()
            
            # Adjust the medicine stock
            medicine = stock.medicine
            medicine.total_quantity += (stock.total_quantity - old_total_quantity)
            medicine.save()
            
            messages.success(request, "Medicine stock updated successfully!")
            return redirect('medicine_stock_list')
    else:
        form = MedicineStockForm(instance=stock)
    return render(request, 'medicine_stock/update-medicine-stock.html', {'form': form})

@login_required
@user_has_access('product_management')
def delete_medicine_stock(request, pk):
    stock = get_object_or_404(MedicineStockModel, pk=pk)
    medicine = stock.medicine
    medicine.total_quantity -= stock.total_quantity
    medicine.save()
    stock.delete()
    messages.success(request, "Medicine stock deleted successfully!")
    return redirect('medicine_stock_list')

def delete_selected_stocks(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_stocks")

        if selected_ids:
            stocks = MedicineStockModel.objects.filter(id__in=selected_ids)

            for stock in stocks:
                medicine = stock.medicine
                if medicine:
                    medicine.total_quantity -= stock.total_quantity
                    medicine.save()

                stock.delete()

            messages.success(request, "Selected medicine stock records deleted successfully.")
        else:
            messages.warning(request, "No medicine stock records selected for deletion.")

    return redirect("medicine_stock_list")

#------Low stock
@login_required
@user_has_access('product_management','low_stocks','billing_management')
def low_stocks(request):
    # Fetch both low stock and out-of-stock data in one query
    low_stock_data = MedicineModel.objects.filter(
        Q(total_quantity__lt=10) | Q(stocks='Out of Stock')
    ).values('id', 'medicine_name', 'pack_size', 'total_quantity', 'unit_sale_price', 'stocks')

    # Create two lists from the fetched data based on 'stocks' value
    low_stock_medicines = [medicine for medicine in low_stock_data if medicine['total_quantity'] < 10]
    out_of_stock_medicines = [medicine for medicine in low_stock_data if medicine['stocks'] == 'Out of Stock']

    if request.GET.get('download') == 'true':
        # Create an Excel workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Out of Stock Medicine List"

        # Write headers
        sheet.append(['medicine_name', 'total_quantity', 'purchase_price'])

        # Write out-of-stock medicine data with two empty columns
        for medicine in out_of_stock_medicines:
            sheet.append([medicine['medicine_name'], '', ''])  # Adding two blank columns

        # Prepare HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="out_of_stock_medicine_list.xlsx"'

        # Save workbook to response
        workbook.save(response)
        return response
    
    context = {
        'low_stock_data': low_stock_medicines,
    }
    return render(request, 'medicine_stock/low-stock.html', context)


#--------Bottle Breakage
@login_required
@user_has_access('product_management')
def add_bottle_breakage(request):
    if request.method == "POST":
        form = BottleBreakageForm(request.POST)
        if form.is_valid():
            bottle_breakage = form.save(commit=False)
            bottle_breakage.created_by = request.user
            
            medicine = bottle_breakage.medicine
            if medicine:
                
                try:
                    if medicine.total_medicine < bottle_breakage.lost_quantity:
                        messages.warning(request, f"Invalid Lost Quantity! Available stock: {medicine.total_medicine}{medicine.pack_units}")
                        return redirect('add_bottle_breakage')
                    else:
                        medicine.total_medicine -= Decimal(bottle_breakage.lost_quantity)
                        medicine.total_quantity -= Decimal(bottle_breakage.lost_quantity) / Decimal(medicine.pack_size)
                        medicine.save()

                except MedicineModel.DoesNotExist:
                    messages.warning(request, "Medicine data not found.")
                    return redirect('add_bottle_breakage')
            
            bottle_breakage.save()
            messages.success(request, "Bottle breakage recorded successfully.")
            return redirect('bottle_breakage_list')
    else:
        form = BottleBreakageForm()
    
    return render(request, "bottle_breakage/add-bottle-breakage.html", {"form": form})

# Update Bottle Breakage Entry
@login_required
@user_has_access('product_management')
def update_bottle_breakage(request, pk):
    bottle_breakage = get_object_or_404(BottleBreakageModel, pk=pk)
    previous_lost_quantity = bottle_breakage.lost_quantity
    
    if request.method == "POST":
        form = BottleBreakageForm(request.POST, instance=bottle_breakage)
        if form.is_valid():
            updated_breakage = form.save(commit=False)
            medicine = updated_breakage.medicine
            
            if medicine:
                try:
                    medicine.total_medicine += previous_lost_quantity
                    medicine.total_quantity += previous_lost_quantity / (medicine.pack_size)
                    
                    if medicine.total_medicine < updated_breakage.lost_quantity:
                        messages.warning(request, f"Invalid Lost Quantity! Available stock: {medicine.total_medicine}{medicine.pack_units}")
                        return redirect('update_bottle_breakage', pk=pk)
                    else:
                        medicine.total_medicine -= updated_breakage.lost_quantity
                        
                        medicine.total_medicine -= Decimal(updated_breakage.lost_quantity)
                        medicine.total_quantity -= Decimal(updated_breakage.lost_quantity) / Decimal(medicine.pack_size)
                        
                        medicine.save()
                except MedicineModel.DoesNotExist:
                    messages.warning(request, "Medicine data not found.")
                    return redirect('update_bottle_breakage', pk=pk)
            
            updated_breakage.save()
            messages.success(request, "Bottle breakage updated successfully.")
            return redirect('bottle_breakage_list')
    else:
        form = BottleBreakageForm(instance=bottle_breakage)
    
    return render(request, "bottle_breakage/update-bottle-breakage.html", {"form": form})


@login_required
@user_has_access('product_management')
def delete_bottle_breakage(request, pk):
    bottle_breakage = get_object_or_404(BottleBreakageModel, pk=pk)
    medicine = bottle_breakage.medicine
    
    if medicine:
        try:
            medicine.total_medicine += bottle_breakage.lost_quantity
            medicine.total_quantity += (bottle_breakage.lost_quantity) / (medicine.pack_size)
            medicine.save()
        except MedicineModel.DoesNotExist:
            messages.warning(request, "Medicine data not found.")
            return redirect('bottle_breakage_list')
    
    bottle_breakage.delete()
    messages.success(request, "Bottle breakage record deleted successfully.")
    return redirect('bottle_breakage_list')

def delete_selected_bottle_breakages(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_bottle_breakages")

        if selected_ids:
            bottle_breakages = BottleBreakageModel.objects.filter(id__in=selected_ids)

            for bottle_breakage in bottle_breakages:
                medicine = bottle_breakage.medicine

                if medicine:
                    try:
                        medicine.total_medicine += bottle_breakage.lost_quantity
                        medicine.total_quantity += bottle_breakage.lost_quantity / medicine.pack_size
                        medicine.save()
                    except MedicineModel.DoesNotExist:
                        messages.warning(request, "Medicine data not found for one or more records.")
                        continue

                bottle_breakage.delete()

            messages.success(request, "Selected bottle breakage records deleted successfully.")
        else:
            messages.warning(request, "No bottle breakage records selected for deletion.")

    return redirect("bottle_breakage_list")

        

# List Bottle Breakages
@login_required
@user_has_access('product_management')
def bottle_breakage_list(request):
    bottle_breakages = BottleBreakageModel.objects.all().order_by('-id')
    return render(request, "bottle_breakage/bottle-breakage-list.html", {"bottle_breakages": bottle_breakages})


#---------Billing Functionalities
def billing_generate():
    return f"ORD-{random.randint(100000, 999999)}"

@login_required
@user_has_access('billing_management')
def billing_list(request):
    billing_status_filter = request.GET.get('status', 'All')
    
    if billing_status_filter == 'All':
        billings = BillingModel.objects.annotate(
            total_items=Count('billing_items'), 
            total_price=Sum('billing_items__total_price'),
            total_medicine_quantity=Sum('billing_items__medicine_quantity')  
        ).order_by('-id')
    elif billing_status_filter == 'Due':
        billings = BillingModel.objects.filter(billing_status=billing_status_filter).annotate(
            total_items=Count('billing_items'),
            total_price=Sum('billing_items__total_price'),
            total_medicine_quantity=Sum('billing_items__medicine_quantity')
        ).order_by('-id')
    else:
        billings = BillingModel.objects.exclude(billing_status='Due').annotate(
            total_items=Count('billing_items'),
            total_price=Sum('billing_items__total_price'),
            total_medicine_quantity=Sum('billing_items__medicine_quantity')
        ).order_by('-id')

    return render(request, 'billings/billing-list.html', {'billings': billings, 'billing_status_filter': billing_status_filter})

@login_required
@user_has_access('billing_management')
def billing_create(request):
    medicines = MedicineModel.objects.all()  # Fetch all medicines
    
    if request.method == "POST":
        billing_form = BillingForm(request.POST)
        
        if billing_form.is_valid():
            billing = billing_form.save(commit=False)
            phone = billing_form.cleaned_data['customer_phone']
            phone_with_code = f"+91{phone}"

            # Check if customer exists
            try:
                customer = CustomerModel.objects.get(customer_phone=phone)
                billing.customer_user = customer
            except CustomerModel.DoesNotExist:
                # Create a new customer
                if len(phone) != 10:
                    messages.warning(request, "Phone number must be 10 digit. Used valid phone number")
                    billing_form.add_error('customer_phone', 'Phone number must be at least 10 digits long.')
                    return render(request, 'billings/add-billing.html', {'billing_form': billing_form, 'medicines': medicines})
                if CustomerModel.objects.filter(customer_phone=phone_with_code).exists():
                    messages.warning(request, "Phone number is already taken!")
                    return render(request, 'billings/update-billing.html', {'billing_form': billing_form, 'medicines': medicines, 'billing': billing })
                customer = CustomerModel.objects.create(
                    customer_name=billing_form.cleaned_data['customer_name'],
                    customer_phone=phone_with_code,
                    customer_email=billing_form.cleaned_data['customer_email'],
                    customer_dob=billing_form.cleaned_data['customer_dob'], 
                    customer_address=billing_form.cleaned_data['customer_address'],
                    created_by=request.user,
                )
                billing.customer_user = customer
                billing.customer_phone = phone_with_code

            # Generate unique billing number
            while True:
                billing_no = billing_generate()
                if not BillingModel.objects.filter(billing_no=billing_no).exists():
                    break

            billing.billing_no = billing_no
            billing.created_by = request.user
            billing.total_amount = Decimal(0)

            # Initialize a flag to track stock issues
            stock_issue = False
            out_of_stock_medicines = []

            # Processing multiple billing items
            medicines_ids = request.POST.getlist('medicine[]')
            quantities = request.POST.getlist('medicine_quantity[]')
            calculation_types = request.POST.getlist('calculation_type[]')

            # First, check stock availability for all medicines
            for i in range(len(medicines_ids)):
                medicine = MedicineModel.objects.get(id=medicines_ids[i])
                medicine_quantity = Decimal(quantities[i])
                
                if calculation_types[i] == 'Pack':
                    if medicine_quantity > medicine.total_quantity:
                        out_of_stock_medicines.append(medicine.medicine_name)
                        stock_issue = True
                
                elif calculation_types[i] == 'Unit':
                    if medicine_quantity > medicine.total_medicine:
                        out_of_stock_medicines.append(medicine.medicine_name)
                        stock_issue = True

            if stock_issue:
                # If any stock issue occurred, show a message and don't save the billing
                medicine_names = ', '.join(out_of_stock_medicines)
                messages.warning(request, f"Stock not available for the following medicine(s): {medicine_names}")
                return render(request, 'billings/add-billing.html', {'billing_form': billing_form, 'medicines': medicines})

            # billing.customer_phone = phone_with_code
            billing.save()
            for i in range(len(medicines_ids)):
                medicine = MedicineModel.objects.get(id=medicines_ids[i])
                medicine_quantity = Decimal(quantities[i])
                calculation_type = calculation_types[i]
                unit_sale_price = Decimal(medicine.unit_sale_price)

                # Check for stock before proceeding with medicine update
                if calculation_type == 'Pack':
                    medicine.total_quantity -= medicine_quantity
                    medicine.total_medicine -= medicine.pack_size * medicine_quantity
                    total_price = (medicine_quantity * unit_sale_price) + medicine.gst_amount
                
                elif calculation_type == 'Unit':
                    medicine.total_medicine -= Decimal(medicine_quantity)
                    medicine.total_quantity -= Decimal(medicine_quantity) / Decimal(medicine.pack_size)
                    total_price = ((Decimal(medicine_quantity) / Decimal(medicine.pack_size)) * unit_sale_price) + medicine.gst_amount
                    
                
                billing.total_amount += total_price
                billing.discount_amount = billing.total_amount * (billing.discount_percentage / Decimal('100'))
                billing.total_amount -= billing.discount_amount
                billing.save()

                # Create BillingItem after saving the BillingModel
                BillingItemModel.objects.create(
                    billing=billing,
                    medicine=medicine,
                    medicine_quantity=medicine_quantity,
                    calculation_type=calculation_type,
                    unit_sale_price=unit_sale_price,
                    gst_amount=medicine.gst_amount,
                    gst_percentage=medicine.gst_percentage,
                    total_price=total_price
                )

                medicine.save()

            generate_invoice(request, billing.id)

            messages.success(request, "Billing created successfully!")
            return redirect('invoice', billing.id)

    else:
        billing_form = BillingForm()
    
    context = {
        'billing_form': billing_form,
        'medicines': medicines,
    }

    return render(request, 'billings/add-billing.html', context)


def get_customer_by_phone(request, phone):
    try:
        customer = CustomerModel.objects.get(customer_phone=phone)
        data = {
            'exists': True,
            'customer_name': customer.customer_name,
            'customer_email': customer.customer_email,
            'customer_dob': customer.customer_dob,
            'customer_address': customer.customer_address,
        }
    except CustomerModel.DoesNotExist:
        data = {'exists': False}
    return JsonResponse(data)

def get_customer_by_phone_autocomplete(request):
    term = request.GET.get('term', '')
    customers = CustomerModel.objects.filter(customer_phone__icontains=term).values_list('customer_phone', flat=True)
    return JsonResponse(list(customers), safe=False)

@login_required
@user_has_access('billing_management')
def billing_update(request, pk):
    billing = get_object_or_404(BillingModel, id=pk)
    medicines = MedicineModel.objects.all()
    billing_items = billing.billing_items.all()
    
    if request.method == "POST":
        billing_form = BillingForm(request.POST, instance=billing)
        
        if billing_form.is_valid():
            updated_billing = billing_form.save(commit=False)
            phone = billing_form.cleaned_data['customer_phone']
            phone_with_code = f"+91{phone}"

            # Check if customer exists
            try:
                customer = CustomerModel.objects.get(customer_phone=phone)
                updated_billing.customer_user = customer
            except CustomerModel.DoesNotExist:
                # Create a new customer
                # Validate phone number length
                if len(phone) < 10:
                    billing_form.add_error('customer_phone', 'Phone number must be at least 10 digits long.')
                    messages.warning(request, "Phone number is less than 10. Used valid phone number")
                    return render(request, 'billings/update-billing.html', {
                        'billing_form': billing_form,
                        'medicines': medicines,
                        'billing': billing,
                        'billing_items': billing_items
                    })
                if CustomerModel.objects.filter(customer_phone=phone_with_code).exists():
                    messages.warning(request, "Phone number is already taken!")
                    return render(request, 'billings/update-billing.html', {'billing_form': billing_form, 'medicines': medicines, 'billing': billing, 'billing_items': billing_items})
                customer = CustomerModel.objects.create(
                    customer_name=billing_form.cleaned_data['customer_name'],
                    customer_phone=phone_with_code,
                    customer_email=billing_form.cleaned_data['customer_email'],
                    customer_dob=billing_form.cleaned_data['customer_dob'],
                    customer_address=billing_form.cleaned_data['customer_address'],
                    created_by=request.user,
                )
                updated_billing.customer_user = customer
            
            # Restore stock before updating items
            for item in billing_items:
                if item.calculation_type == 'Pack':
                    item.medicine.total_quantity += item.medicine_quantity
                    item.medicine.total_medicine += item.medicine.pack_size * item.medicine_quantity
                elif item.calculation_type == 'Unit':
                    item.medicine.total_medicine += Decimal(item.medicine_quantity)
                    item.medicine.total_quantity += Decimal(item.medicine_quantity) / Decimal(item.medicine.pack_size)
                item.medicine.save()
                item.delete()
            
            updated_billing.total_amount = Decimal(0)
            medicines_ids = request.POST.getlist('medicine[]')
            quantities = request.POST.getlist('medicine_quantity[]')
            calculation_types = request.POST.getlist('calculation_type[]')
            
            stock_issue = False
            out_of_stock_medicines = []
            
            # Check stock availability
            for i in range(len(medicines_ids)):
                medicine = MedicineModel.objects.get(id=medicines_ids[i])
                medicine_quantity = Decimal(quantities[i])
                calculation_type = calculation_types[i]
                
                if calculation_type == 'Pack' and medicine_quantity > medicine.total_quantity:
                    out_of_stock_medicines.append(medicine.medicine_name)
                    stock_issue = True
                elif calculation_type == 'Unit' and medicine_quantity > medicine.total_medicine:
                    out_of_stock_medicines.append(medicine.medicine_name)
                    stock_issue = True
            
            if stock_issue:
                messages.warning(request, f"Stock not available for: {', '.join(out_of_stock_medicines)}")
                return render(request, 'billings/update-billing.html', {'billing_form': billing_form, 'medicines': medicines, 'billing': billing, 'billing_items': billing_items})
            
            # Save updated billing
            updated_billing.save()
            
            # Process new billing items
            for i in range(len(medicines_ids)):
                medicine = MedicineModel.objects.get(id=medicines_ids[i])
                medicine_quantity = Decimal(quantities[i])
                calculation_type = calculation_types[i]
                unit_sale_price = Decimal(medicine.unit_sale_price)
                
                if calculation_type == 'Pack':
                    medicine.total_quantity -= medicine_quantity
                    medicine.total_medicine -= medicine.pack_size * medicine_quantity
                    total_price = (medicine_quantity * unit_sale_price) + medicine.gst_amount
                elif calculation_type == 'Unit':
                    medicine.total_medicine -= Decimal(medicine_quantity)
                    medicine.total_quantity -= Decimal(medicine_quantity) / Decimal(medicine.pack_size)
                    total_price = ((Decimal(medicine_quantity) / Decimal(medicine.pack_size)) * unit_sale_price) + medicine.gst_amount
                
                BillingItemModel.objects.create(
                    billing=updated_billing,
                    medicine=medicine,
                    medicine_quantity=medicine_quantity,
                    calculation_type=calculation_type,
                    unit_sale_price=unit_sale_price,
                    gst_amount=medicine.gst_amount,
                    gst_percentage=medicine.gst_percentage,
                    total_price=total_price
                )
                
                medicine.save()
                updated_billing.total_amount += total_price
            

            updated_billing.discount_amount = updated_billing.total_amount * (updated_billing.discount_percentage / Decimal('100'))
            updated_billing.total_amount -= updated_billing.discount_amount
            updated_billing.save()
            
            messages.success(request, "Billing updated successfully!")
            return redirect('billing_list')
    
    else:
        billing_form = BillingForm(instance=billing)
    
    return render(request, 'billings/update-billing.html', {
        'billing_form': billing_form,
        'medicines': medicines,
        'billing': billing,
        'billing_items': billing_items
    })
    
@login_required
@user_has_access('billing_management')
def billing_delete(request, pk):
    billing = get_object_or_404(BillingModel, pk=pk)

    for billing_item in billing.billing_items.all():
        medicine = billing_item.medicine
        
        if billing_item.calculation_type == 'Pack':
            medicine.total_quantity += billing_item.medicine_quantity
            medicine.total_medicine += Decimal(billing_item.medicine_quantity) / Decimal(medicine.pack_size)
        elif billing_item.calculation_type == 'Unit':
            medicine.total_medicine += Decimal(billing_item.medicine_quantity)
            medicine.total_quantity += Decimal(billing_item.medicine_quantity) / Decimal(medicine.pack_size)
        medicine.save()

    billing.delete()
    messages.success(request, "Billing deleted successfully!")
    return redirect('billing_list')

def delete_selected_billings(request):
    if request.method == "POST":
        selected_ids = request.POST.getlist("selected_billings")

        if selected_ids:
            billings = BillingModel.objects.filter(id__in=selected_ids)

            for billing in billings:
                for billing_item in billing.billing_items.all():
                    medicine = billing_item.medicine
                    
                    if billing_item.calculation_type == "Pack":
                        medicine.total_quantity += billing_item.medicine_quantity
                        medicine.total_medicine += Decimal(billing_item.medicine_quantity) / Decimal(medicine.pack_size)
                    elif billing_item.calculation_type == "Unit":
                        medicine.total_medicine += Decimal(billing_item.medicine_quantity)
                        medicine.total_quantity += Decimal(billing_item.medicine_quantity) / Decimal(medicine.pack_size)
                    
                    medicine.save()

                billing.delete()

            messages.success(request, "Selected billings deleted successfully.")
        else:
            messages.warning(request, "No billings selected for deletion.")

    return redirect("billing_list")

@login_required
@user_has_access('billing_management')
def invoice_list(request):
    billing_status_filter = request.GET.get('status', 'All')
    
    if billing_status_filter == 'All': 
        billings = BillingModel.objects.annotate(
            total_items=Count('billing_items'),
            total_price=Sum('billing_items__total_price'),  
            total_medicine_quantity=Sum('billing_items__medicine_quantity')  
        ).order_by('-id')
    else:
        billings = BillingModel.objects.filter(billing_status=billing_status_filter).annotate(
            total_items=Count('billing_items'),
            total_price=Sum('billing_items__total_price'),  
            total_medicine_quantity=Sum('billing_items__medicine_quantity')  
        ).order_by('-id')
    
    
    return render(request, 'invoices/invoice-list.html', {'billings': billings,'billing_status_filter': billing_status_filter})

@login_required
@user_has_access('billing_management')
def invoice(request, billing_id):
    # Fetch the billing with the given billing_id and related billing items
    billing = get_object_or_404(BillingModel, id=billing_id)
    billing_items = BillingItemModel.objects.filter(billing=billing)
    subtotal = sum(item.total_price for item in billing_items)

    # Pass the billing and billing items to the template
    context = {
        'billing': billing,
        'billing_items': billing_items,
        'subtotal': subtotal,
    }
    return render(request, 'invoices/invoice.html', context)


def generate_invoice(request, billing_id):
    invoice = BillingModel.objects.get(id=billing_id)
    billing_items = BillingItemModel.objects.filter(billing=invoice)
    subtotal = sum(item.total_price for item in billing_items)

    context = {
        'billing': invoice,
        'billing_items': billing_items,
        'subtotal': subtotal,
    }
    html_string = render_to_string('invoices/invoice_template.html',context)

    pdf = HTML(string=html_string).write_pdf()

    invoice_folder = os.path.join(settings.MEDIA_ROOT, 'invoices')
    os.makedirs(invoice_folder, exist_ok=True)

    # Define the file path
    file_path = os.path.join(invoice_folder, f'invoice_{invoice.billing_no}.pdf')

    # Save the PDF to the file
    with open(file_path, 'wb') as f:
        f.write(pdf)

    # Optionally store the file path in the database (optional)
    invoice.pdf_file = os.path.join('invoices', f'invoice_{invoice.billing_no}.pdf')
    invoice.save()

    # Return the PDF in the response without triggering a download
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=invoice_{invoice.billing_no}.pdf'  # This opens the PDF in the browser

    return response

@login_required
@user_has_access('product_management','inventory_report')
def inventory_report(request):
    # Query the MedicineModel with the related stock and sales data
    inventory_report = MedicineModel.objects.annotate(
        total_stock=Sum('medicinestocks__total_quantity'),
        total_sales=Sum('medicine_billings__medicine_quantity'),
        total_loss=Sum('breakages__lost_quantity')
    ).values(
        'medicine_name', 
        'total_stock', 
        'total_sales', 
        'total_loss', 
        'unit_sale_price'
    )

    # For downloading the report in CSV format
    if request.GET.get('download') == 'true':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'
        writer = csv.writer(response)

        # Write headers to CSV file
        writer.writerow(['Medicine Name', 'Total Stock', 'Total Sales', 'Total Loss', 'Unit Sale Price'])

        # Write data to CSV file
        for item in inventory_report:
            writer.writerow([item['medicine_name'], item['total_stock'], item['total_sales'], item['total_loss'], item['unit_sale_price']])

        return response

    return render(request, 'reports/inventory-report.html', {'inventory_report': inventory_report})

@login_required
@user_has_access('product_management','wastage_report')
def wastage_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_employee_id = request.GET.get('employee_id')
    selected_medicine_id = request.GET.get('medicine_name')

    # Default query to fetch BottleBreakageModel objects with filtering logic
    wastage_report = BottleBreakageModel.objects.select_related(
        'medicine', 
        'responsible_employee', 
    ).values(
        'medicine__medicine_name', 
        'lost_quantity', 
        'responsible_employee__employee_user__username',
        'date_time',
        'id',
    )

    # Apply date filter if both start_date and end_date are provided
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            wastage_report = wastage_report.filter(date_time__range=[start_date, end_date])
        except ValueError:
            wastage_report = wastage_report.none()

    # Apply employee filter if selected
    if selected_employee_id:
        wastage_report = wastage_report.filter(responsible_employee__id=selected_employee_id)

    # Apply medicine filter if selected
    if selected_medicine_id:
        wastage_report = wastage_report.filter(medicine__id=selected_medicine_id)

    # Fetch employees and medicines used in BottleBreakageModel to pass to the template
    employees = BottleBreakageModel.objects.values('responsible_employee__id', 'responsible_employee__employee_user__username').distinct()
    medicines = BottleBreakageModel.objects.values('medicine__id', 'medicine__medicine_name').distinct()

    # Export to CSV functionality
    if request.GET.get('download') == 'true':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="wastage_report.csv"'
        writer = csv.writer(response)

        # Write headers to CSV file
        writer.writerow(['Medicine Name', 'Lost Quantity', 'Responsible Employee', 'Date'])

        # Write data to CSV file
        for item in wastage_report:
            writer.writerow([item['medicine__medicine_name'], item['lost_quantity'], item['responsible_employee__employee_user__username'], item['date_time']])

        return response

    return render(request, 'reports/wastage-report.html', {
        'wastage_report': wastage_report,
        'start_date': start_date,
        'end_date': end_date,
        'employees': employees,
        'medicines': medicines,
        'selected_employee_id': selected_employee_id,
        'selected_medicine_id': selected_medicine_id,
    })


@login_required  
@user_has_access('billing_report','billing_management')
def billing_trends_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    billing_trends = BillingItemModel.objects.annotate(
        medicine_name=F('medicine__medicine_name'),
        billing_date=F('billing__billing_date'), 
        total_sales=F('medicine_quantity'),
        total_revenue=F('total_price'),
    ).values(
        'medicine_name',
        'billing_date',
        'total_sales',
        'total_revenue',
    ).order_by('-billing_date')

    # Filter by date range if provided
    if start_date and end_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Extend end_date to include the entire day
            end_datetime = datetime.combine(end_date, datetime.max.time())

            billing_trends = billing_trends.filter(billing__billing_date__range=[start_date, end_datetime])
        except ValueError:
            # Handle invalid date format
            billing_trends = billing_trends.none()

    # For CSV download
    if request.GET.get('download') == 'true':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="billing_trends_report.csv"'
        writer = csv.writer(response)

        # Write headers to CSV file
        writer.writerow(['Medicine Name', 'Total Qty', 'Total Revenue','Timestamp'])

        # Write data to CSV file
        for item in billing_trends:
            writer.writerow([item['medicine_name'], item['total_sales'], item['total_revenue'], item['billing_date']])

        return response

    return render(request, 'reports/billing-trends-report.html', {
        'billing_trends': billing_trends,
        'start_date': start_date,
        'end_date': end_date,
    })
    


#--------send invoice
def send_invoice_email(request, billing_id):
    billing = get_object_or_404(BillingModel, id=billing_id)
    
    subject = "Your Billing Invoice"
    message = "Dear Customer, Please find attached your invoice."

    to_email = billing.customer_email
    
    from_email = settings.EMAIL_HOST_USER
    email = EmailMessage(subject, message, from_email, [to_email])

    if billing.pdf_file:
        with open(billing.pdf_file.path, 'rb') as pdf:
            email.attach('invoice.pdf', pdf.read(), 'application/pdf')

        email.send()
        messages.success(request, "Invoice sent successfully to the customer.")
    else:
        messages.error(request, "Invoice file not found.")
    return redirect('invoice_list') 


def send_invoice_via_whatsapp(request, billing_id):
    billing = get_object_or_404(BillingModel, id=billing_id)

    if not billing.pdf_file:
        return HttpResponse("Invoice PDF not available", status=404)

    # Ensure the phone number has the correct country code
    customer_phone = billing.customer_phone
    if not customer_phone.startswith("+"):
        customer_phone = f"+88{customer_phone}"  # Adjust based on your country

    pdf_file_url = f'http://inventory.mrshakil.com{billing.pdf_file.url}'

    # Encode the message properly
    message = f"Hi, please find my invoice here: {pdf_file_url}"
    encoded_message = quote(message)  # Encode special characters properly

    # Construct the WhatsApp URL
    whatsapp_url = f"https://wa.me/{customer_phone}?text={encoded_message}"

    return redirect(whatsapp_url)